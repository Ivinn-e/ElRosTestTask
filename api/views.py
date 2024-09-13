from .serializers import CountrySerializer, ManufacturerSerializer, CarSerializer, CommentsSerializer
from .models import Country, Manufacturer, Car, Comments
import io
from django.http import HttpResponse
from openpyxl import Workbook
import csv
from rest_framework.views import APIView
import xlrd
from rest_framework.generics import (ListAPIView, CreateAPIView, RetrieveUpdateDestroyAPIView)
from openpyxl.utils import get_column_letter
from io import StringIO 
from rest_framework.authentication import SessionAuthentication
from rest_framework.permissions import IsAuthenticated


# Endpoint для просмотра записи модели Country
class CountryListAPIView(ListAPIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)
    queryset = Country.objects.all()
    serializer_class = CountrySerializer


#Endpoint для создания записи модели Country
class CountryCreateAPIView(CreateAPIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,) 
    queryset = Country.objects.all()
    serializer_class = CountrySerializer

#Endpoint для редактирования, обновления и удаления записи модели Country
class CountryRetrieveUpdateDestroyAPIView(RetrieveUpdateDestroyAPIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)
    queryset = Country.objects.all()
    serializer_class = CountrySerializer

# Endpoint для просмотра записи модели Manufacturer
class ManufacturerListAPIView(ListAPIView):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer

#Endpoint для создания записи модели Manufacturer
class ManufacturerCreateAPIView(CreateAPIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer

#Endpoint для редактирования, обновления и удаления записи модели Manufacturer
class ManufacturerRetrieveUpdateDestroyAPIView(RetrieveUpdateDestroyAPIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer

# Endpoint для просмотра записи модели Car
class CarListAPIView(ListAPIView):
    queryset = Car.objects.all()
    serializer_class = CarSerializer

#Endpoint для создания записи модели Car
class CarCreateAPIView(CreateAPIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)
    queryset = Car.objects.all()
    serializer_class = CarSerializer

#Endpoint для редактирования, обновления и удаления записи модели Car
class CarRetrieveUpdateDestroyAPIView(RetrieveUpdateDestroyAPIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)
    queryset = Car.objects.all()
    serializer_class = CarSerializer

# Endpoint для просмотра записи модели Comments
class CommentsListAPIView(ListAPIView):
    queryset = Comments.objects.all()
    serializer_class = CommentsSerializer

#Endpoint для создания записи модели Comments
class CommentsCreateAPIView(CreateAPIView):
    queryset = Comments.objects.all()
    serializer_class = CommentsSerializer

#Endpoint для редактирования, обновления и удаления записи модели Comments
class CommentsRetrieveUpdateDestroyAPIView(RetrieveUpdateDestroyAPIView):
    permission_classes = (IsAuthenticated,)
    authentication_classes = (SessionAuthentication,)
    queryset = Comments.objects.all()
    serializer_class = CommentsSerializer

class ExportCountryDataXlsxView(APIView):
    def get(self, request):

        countries = Country.objects.all()
        serializer = CountrySerializer(countries, many=True)

        wb = Workbook()
        ws = wb.active
        headers = ['ID', 'Название страны', 'Производители']
        ws.append(headers)

        for country_data in serializer.data:
            manufacturers = ', '.join(country_data['manufacturers'])
            ws.append([country_data['id'], country_data['name'], manufacturers])

        file_buffer = io.BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)

        response = HttpResponse(file_buffer.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=country_export.xlsx'
        return response

class ExportManufacturerDataXlsxView(APIView):
    def get(self, request):
        manufacturers = Manufacturer.objects.all()
        serializer = ManufacturerSerializer(manufacturers, many=True)

        wb = Workbook()
        ws = wb.active
        headers = ['ID', 'Название производителя', 'Страна', 'Автомобили', 'Количество комментариев к автомобилям']
        ws.append(headers)

        for manufacturer_data in serializer.data:
            country_name = manufacturer_data['country'] if 'country' in manufacturer_data else 'N/A'
            cars = ','.join(car['name'] for car in manufacturer_data['cars'])
            ws.append([manufacturer_data['id'], manufacturer_data['name'], country_name, cars, sum(car['comment_count'] for car in manufacturer_data['cars'])])

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True
            ws.column_dimensions[get_column_letter(col)].bestFit = True

        file_buffer = io.BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)

        response = HttpResponse(file_buffer.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=manufacturer_export.xlsx'
        return response
    
class ExportCarDataXlsxView(APIView):
    def get(self, request):
        cars = Car.objects.all()
        serializer = CarSerializer(cars, many=True)

        wb = Workbook()
        ws = wb.active

        headers = ['ID', 'Название автомобиля', 'Производитель', 'Год начала выпуска', 'Год окончания выпуска', 'Количество комментариев']
        ws.append(headers)

        for car_data in serializer.data:
            ws.append([car_data['id'],car_data['name'],car_data['manufacturer'],  car_data['yearOnRelease'],car_data['yearOfGraduation'],car_data['comment_count']])

        file_buffer = io.BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)

        response = HttpResponse(file_buffer.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=car_export.xlsx'
        return response
    
class ExportCommentsDataXlsxView(APIView):
    def get(self, request):
        comments = Comments.objects.all()
        serializer = CommentsSerializer(comments, many=True)

        wb = Workbook()
        ws = wb.active

        headers = ['ID', 'Email', 'Автомобиль', 'Комментарий']
        ws.append(headers)

        for comment_data in serializer.data:
            ws.append([comment_data['id'], comment_data['email'], comment_data['car'], comment_data['comment'], ])

        file_buffer = io.BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)

        response = HttpResponse(file_buffer.read(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=comments_export.xlsx'
        return response
    
class ExportCountryDataCSVView(APIView):
    def get(self, request):
        countries = Country.objects.all()

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="countries.csv"'

        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer)

        writer.writerow(['ID', 'Название страны'])

        for country in countries:
            writer.writerow([country.id, country.name])

        csv_data = csv_buffer.getvalue().encode('utf-8')

        response.write(csv_data)
        return response
    
class ExportManufaturerDataCSVView(APIView):
    def get(self, request):
        manufacturers = Manufacturer.objects.all()

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="manufacturers.csv"'

        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer)

        writer.writerow(['ID', 'Производитель', 'Страна'])

        for manufacturer in manufacturers:
            writer.writerow([manufacturer.id,  manufacturer.name,  manufacturer.country.name ])

        csv_data = csv_buffer.getvalue().encode('utf-8')

        response.write(csv_data)

        return response
    
class ExportCarDataCSVView(APIView):
    def get(self, request):
        cars = Car.objects.all()
        serializer = CarSerializer(cars, many=True)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="cars.csv"'

        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer)

        writer.writerow(['ID', 'Автомобиль', 'Производитель', 'Год начала выпуска', 'Год окончания выпуска', 'Количество комментариев'])

        for car_data in serializer.data:
            writer.writerow([car_data['id'], car_data['name'], car_data['manufacturer']['name'], car_data['yearOnRelease'], car_data['yearOfGraduation'], car_data['comment_count'],])

        csv_data = csv_buffer.getvalue().encode('utf-8')
        response.write(csv_data)
        return response
    
class ExportCommentDataCSVView(APIView):
    def get(self, request):
        comments = Comments.objects.all()
        serializer = CommentsSerializer(comments, many=True)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="comments.csv"'

        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer)

        writer.writerow(['ID', 'Email', 'Автомобиль', 'Комментарий', 'Производитель'])

        for comment_data in serializer.data:
            writer.writerow([comment_data['id'], comment_data['email'], comment_data['car'], comment_data['comment'], comment_data['manufacturer'],])

        csv_data = csv_buffer.getvalue().encode('utf-8')
        response.write(csv_data)
        return response